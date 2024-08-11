from datetime import datetime
import openpyxl

class ExcelHandler:
    def __init__(self, filename='uart_data.xlsx'):
        self.filename = filename
        self.load_workbook()

    def load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filename, data_only=True)
            self.worksheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.append(['Time', 'Temperature', 'Humidity', 'Pressure', 'Altitude', 'Velocity', 'Latitude', 'Longitude'])

    def save_data(self, timestamp, data_points):
        self.worksheet.append([timestamp.strftime('%Y-%m-%d %H:%M:%S')] + data_points)
        self.workbook.save(self.filename)

    def read_data(self):
        rows = list(self.worksheet.iter_rows(min_row=2, values_only=True))
        parsed_rows = []

        for row in rows:
            try:
                timestamp_str = row[0]
                if isinstance(timestamp_str, float):
                # Skip rows with invalid date values
                    continue
                timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                parsed_row = (timestamp, *row[1:])
                parsed_rows.append(parsed_row)
            except (TypeError, ValueError) as e:
                print(f"Error parsing row {row}: {e}")
                continue

        return parsed_rows

        


handler = ExcelHandler()
data= handler.read_data()
print (data)


