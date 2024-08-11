import time
from datetime import datetime
import openpyxl
import serial
import pandas as pd

# ExcelHandler class definition
class ExcelHandler:
    def __init__(self, filename='uart_data.xlsx'):
        self.filename = filename
        self.load_workbook()

    def read_last_row(self):
        df = pd.read_excel(self.filename)
        if not df.empty:
            return df.iloc[-1].tolist()
        return None

    def load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filename, data_only=True)
            self.worksheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.append(['Time', 'Temperature', 'Humidity', 'Pressure', 'Altitude', 'Velocity', 'Latitude', 'Longitude'])

    def save_data(self, timestamp, data_points):
        self.worksheet.append([timestamp.strftime('%Y-%m-%d %H:%M:%S')] + data_points)
        self.workbook.save(self.filename)

    def read_data(self):
        rows = list(self.worksheet.iter_rows(min_row=2, values_only=True))
        parsed_rows = []

        for row in rows:
            try:
                timestamp_str = row[0]
                if isinstance(timestamp_str, float):
                    continue
                timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                parsed_row = (timestamp, *row[1:])
                parsed_rows.append(parsed_row)
            except (TypeError, ValueError) as e:
                print(f"Error parsing row {row}: {e}")
                continue

        return parsed_rows

# UARTReader class definition
class UARTReader:
    def __init__(self, port='COM4', baudrate=115200):
        self.ser = None
        self.connect(port, baudrate)

    def connect(self, port, baudrate):
        try:
            self.ser = serial.Serial(port, baudrate, timeout=1)
            print(f"Connected to {port} at {baudrate} baud.")
        except serial.SerialException as e:
            print(f"Error opening serial port: {e}")

    def read_line(self):
        if self.ser and self.ser.in_waiting:
            line = self.ser.readline().decode('utf-8').strip()
            print(f"Received line: {line}")  # Debug statement
            try:
                data_values = line.split(',')
                required_fields = ['Temperature', 'Humidity', 'Pressure', 'Altitude', 'Velocity', 'Latitude', 'Longitude']
                numeric_values = []
                for field in required_fields:
                    matched = False
                    for value in data_values:
                        if value.strip().startswith(field):
                            try:
                                numeric_value = float(value.split(':')[1].strip())
                                numeric_values.append(numeric_value)
                                matched = True
                                break
                            except (ValueError, IndexError) as e:
                                print(f"Error parsing value: {value}. {e}")
                                return None
                    if not matched:
                        print(f"Field {field} not found in received line.")
                        return None
                if len(numeric_values) == len(required_fields):
                    return numeric_values
                else:
                    print("Incomplete data received.")
                    return None
            except ValueError as e:
                print(f"Error parsing line: {line}. {e}")
                return None
        return None
    
    def close(self):
        if self.ser:
            self.ser.close()
            print(f"Closed serial port.")

# Main script to read UART data and save to Excel
def main():
    uart_reader = UARTReader(port='COM4', baudrate=115200)
    excel_handler = ExcelHandler(filename='uart_data.xlsx')

    while True:
        data_points = uart_reader.read_line()
        if data_points:
            timestamp = datetime.now()
            excel_handler.save_data(timestamp, data_points)
            print(f"Data saved: {timestamp} {data_points}")
        time.sleep(1)  # update every 1 second

if __name__ == "__main__":
    main()
